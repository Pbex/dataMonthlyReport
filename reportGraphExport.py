# last modified by Pb in 2020.9.12
# bugfix: chinese string is not allow to show in pyplot graph

import openpyxl as xl
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.font_manager as fm
from pylab import mpl

# used for Chinese string display, pyplot not support Chinese decode in default
# font = {'family': 'SimHei', "size": 24}
# matplotlib.rc('font', **font)
# my_font_title = fm.FontProperties(family='SimHei', size=24)
# my_font_x_label = fm.FontProperties(family='SimHei', size=20)

# used for Chinese string display, pyplot not support Chinese decode in default
mpl.rcParams['font.sans-serif'] = ['SimHei']  # setting all font to use Chinese

# define the struct that used for drawing graph
class GraphClass:
    def __init__(self):
        self.title = ''
        self.data_names = []
        self.data = []
        self.x_label = ''
        self.y_label = ''


# monthly login sum arr
monthlyActiveArr = GraphClass()


def readDataFromXlsx():
    data_table = xl.load_workbook('data//电商数据.xlsx')
    ws = data_table.worksheets[0]
    data_name_temp_arr = []
    data_temp_arr = []
    for r in range(2, 5):
        data_name_temp_arr.append(ws.cell(row=r, column=1).value)
        data_temp_arr.append(ws.cell(row=r, column=2).value)
    monthlyActiveArr.data_names.append(data_name_temp_arr)
    monthlyActiveArr.data.append(data_temp_arr)
    monthlyActiveArr.title = ws.cell(row=1, column=4).value  # graph title
    scale_ls = range(len(monthlyActiveArr.data_names[0]))  # x_label size

    plt.bar(scale_ls, monthlyActiveArr.data[0])
    plt.xticks(scale_ls, monthlyActiveArr.data_names[0])
    plt.title(monthlyActiveArr.title)
    plt.figure()
    plt.show()


readDataFromXlsx()
# end up with pyplot.show() function to show tables
plt.show()
